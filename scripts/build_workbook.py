"""
build_workbook.py
=================
Nexa Parts Supply - Excel Master Database System
Portfolio Demo | Synthetic Data Only

Reads processed CSV/report files and builds the professional
Excel master workbook: data/processed/nexa_parts_master_database.xlsx

Sheets:
  1. README
  2. Master Products
  3. Vendor Imports
  4. Duplicate Review
  5. Data Issues
  6. Import Log
  7. Category Summary
  8. Inventory Dashboard

Usage:
    python scripts/build_workbook.py
"""

from pathlib import Path
from datetime import datetime

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT      = Path(__file__).resolve().parent.parent
PROC_DIR  = ROOT / "data" / "processed"
RPT_DIR   = ROOT / "data" / "reports"
OUT_FILE  = PROC_DIR / "nexa_parts_master_database.xlsx"

# ---------------------------------------------------------------------------
# Colour Palette
# ---------------------------------------------------------------------------
# Primary brand colours
C_NAVY      = "1A2B4A"   # dark navy header
C_BLUE_MID  = "2D5AA0"   # mid blue
C_BLUE_LT   = "4A90D9"   # light blue accent
C_TEAL      = "0E7C7B"   # teal accent
C_GOLD      = "D4A017"   # gold/amber
C_ORANGE    = "E07B39"   # orange warning
C_RED       = "C0392B"   # red error
C_GREEN     = "27AE60"   # green ok
C_PURPLE    = "6C3483"   # purple

# Background / neutral
C_WHITE     = "FFFFFF"
C_OFFWHITE  = "F7F9FC"
C_LIGHT_GRAY= "EEF2F7"
C_MID_GRAY  = "CDD6E0"
C_DARK_GRAY = "7F8C8D"
C_BLACK     = "1C1C1C"

# Row alternating
C_ROW_ALT   = "EEF4FB"   # very light blue alternating row
C_ROW_NORM  = C_WHITE

# Issue severity
C_HIGH_BG   = "FDECEA"
C_HIGH_FG   = "C0392B"
C_MED_BG    = "FEF9E7"
C_MED_FG    = "B7770D"
C_LOW_BG    = "EAF8F0"
C_LOW_FG    = "1D8348"

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def hex_fill(hex_color, fill_type="solid"):
    return PatternFill(fill_type=fill_type, fgColor=hex_color)


def bold_font(size=10, color=C_WHITE, name="Calibri"):
    return Font(name=name, bold=True, size=size, color=color)


def normal_font(size=10, color=C_BLACK, name="Calibri"):
    return Font(name=name, size=size, color=color)


def thin_border():
    s = Side(style="thin", color=C_MID_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)


def thick_bottom():
    thick = Side(style="medium", color=C_NAVY)
    return Border(bottom=thick)


def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def left_align(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def right_align():
    return Alignment(horizontal="right", vertical="center")


# ---------------------------------------------------------------------------
# Apply header row to a worksheet
# ---------------------------------------------------------------------------

def apply_header_row(ws, row_num, headers, col_start=1,
                     bg=C_NAVY, fg=C_WHITE, size=10, height=22):
    ws.row_dimensions[row_num].height = height
    for i, h in enumerate(headers, start=col_start):
        cell = ws.cell(row=row_num, column=i, value=h)
        cell.font      = bold_font(size=size, color=fg)
        cell.fill      = hex_fill(bg)
        cell.alignment = center_align(wrap=True)
        cell.border    = thin_border()


# ---------------------------------------------------------------------------
# Write DataFrame to sheet with styled headers & alternating rows
# ---------------------------------------------------------------------------

def df_to_sheet(ws, df, start_row=1, header_bg=C_NAVY, header_fg=C_WHITE,
                alt_row_fill=True, freeze_row=True, apply_filter=True,
                col_widths=None, number_formats=None):
    """
    Write df to ws starting at start_row.
    Returns the last row written.
    """
    headers = list(df.columns)
    apply_header_row(ws, start_row, headers, bg=header_bg, fg=header_fg)

    num_formats = number_formats or {}

    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        is_alt = (r_idx - start_row) % 2 == 0
        row_fill = hex_fill(C_ROW_ALT) if (alt_row_fill and is_alt) else hex_fill(C_ROW_NORM)
        ws.row_dimensions[r_idx].height = 16

        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font      = normal_font()
            cell.fill      = row_fill
            cell.alignment = left_align()
            cell.border    = thin_border()
            col_name = headers[c_idx - 1]
            if col_name in num_formats:
                cell.number_format = num_formats[col_name]

    # Set column widths
    if col_widths:
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
    else:
        # Auto-width (capped)
        for i, col in enumerate(headers, start=1):
            col_letter = get_column_letter(i)
            max_len = max(len(str(col)), df[col].astype(str).str.len().max() if len(df) > 0 else 0)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)

    # Freeze pane
    if freeze_row:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Auto-filter
    if apply_filter and len(df) > 0:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A{start_row}:{last_col}{start_row}"

    return start_row + len(df)


# ---------------------------------------------------------------------------
# Sheet 1: README
# ---------------------------------------------------------------------------

def build_readme_sheet(ws):
    ws.title = "README"
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 72

    # Title banner
    ws.row_dimensions[2].height = 40
    title_cell = ws.cell(row=2, column=2, value="Nexa Parts Supply")
    title_cell.font      = Font(name="Calibri", bold=True, size=22, color=C_WHITE)
    title_cell.fill      = hex_fill(C_NAVY)
    title_cell.alignment = center_align()

    sub_cell = ws.cell(row=2, column=3, value="Excel Master Database System — Portfolio Demo")
    sub_cell.font      = Font(name="Calibri", bold=True, size=14, color=C_GOLD)
    sub_cell.fill      = hex_fill(C_NAVY)
    sub_cell.alignment = left_align()

    # Disclaimer
    ws.row_dimensions[3].height = 36
    disc_cell = ws.cell(row=3, column=2,
        value="⚠  DISCLAIMER")
    disc_cell.font      = bold_font(size=10, color=C_HIGH_FG)
    disc_cell.fill      = hex_fill(C_HIGH_BG)
    disc_cell.alignment = center_align(wrap=True)

    disc_text = ws.cell(row=3, column=3,
        value=("This is a portfolio demo built with synthetic data. It is designed to reflect realistic "
               "business workflows without using real customer, vendor, company, or confidential data. "
               "All company names, vendor names, product names, SKUs, prices, and contact details are "
               "entirely fictional and generated for demonstration purposes only."))
    disc_text.font      = normal_font(size=9, color=C_HIGH_FG)
    disc_text.fill      = hex_fill(C_HIGH_BG)
    disc_text.alignment = left_align(wrap=True)

    def section_header(row, text):
        ws.row_dimensions[row].height = 20
        hdr = ws.cell(row=row, column=2, value=text)
        hdr.font      = bold_font(size=11, color=C_WHITE)
        hdr.fill      = hex_fill(C_BLUE_MID)
        hdr.alignment = left_align()
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)

    def info_row(row, label, value, label_bg=C_LIGHT_GRAY):
        ws.row_dimensions[row].height = 18
        lbl = ws.cell(row=row, column=2, value=label)
        lbl.font      = bold_font(size=10, color=C_NAVY)
        lbl.fill      = hex_fill(label_bg)
        lbl.alignment = left_align()
        val = ws.cell(row=row, column=3, value=value)
        val.font      = normal_font(size=10)
        val.fill      = hex_fill(C_OFFWHITE)
        val.alignment = left_align(wrap=True)

    r = 5
    section_header(r, "📋  About This Workbook"); r += 1
    info_row(r, "Project",       "Excel Master Database System — Nexa Parts Supply"); r += 1
    info_row(r, "Purpose",       "Portfolio demo: automated vendor data cleaning, merging, and reporting"); r += 1
    info_row(r, "Generated",     datetime.now().strftime("%Y-%m-%d %H:%M:%S")); r += 1
    info_row(r, "Data Type",     "Entirely synthetic — no real vendor or product data"); r += 1

    r += 1
    section_header(r, "📄  Sheet Guide"); r += 1
    sheets_info = [
        ("Master Products",  "Cleaned, deduplicated product master with margin, stock value & status"),
        ("Vendor Imports",   "All raw import records with source tracking and canonical columns"),
        ("Duplicate Review", "SKU groups with more than one record across vendors"),
        ("Data Issues",      "Detected quality issues with severity ratings (HIGH/MEDIUM/LOW)"),
        ("Import Log",       "Per-vendor import run summary with row counts and issue tallies"),
        ("Category Summary", "Aggregated product counts, values, and stock by category and vendor"),
        ("Inventory Dashboard", "KPI cards and summary tables for executive overview"),
    ]
    for sheet, desc in sheets_info:
        info_row(r, sheet, desc); r += 1

    r += 1
    section_header(r, "✅  Data Quality Rules Applied"); r += 1
    rules = [
        ("SKU Cleaning",        "Whitespace removed, normalised to uppercase"),
        ("Price Cleaning",      "Currency symbols stripped, blanks/zeros/negatives flagged"),
        ("Quantity Cleaning",   "Non-numeric values (unknown, out, five) flagged as issues"),
        ("Status Normalisation","Active/Discontinued mapped from 15+ source variants"),
        ("Category Mapping",    "20+ raw category labels mapped to 12 canonical categories"),
        ("Unit Mapping",        "10+ unit variants mapped to 9 canonical UoM values"),
        ("Duplicate Detection", "SKU-based cross-vendor deduplication with best-record selection"),
        ("Severity Rating",     "Issues rated HIGH/MEDIUM/LOW based on business impact"),
    ]
    for rule, desc in rules:
        info_row(r, rule, desc); r += 1

    r += 1
    section_header(r, "🛠  Tech Stack"); r += 1
    stack = [
        ("Language",    "Python 3.x"),
        ("Libraries",   "pandas, openpyxl"),
        ("Data",        "100% synthetic — generated by generate_vendor_data.py"),
        ("Workbook",    "Built programmatically by build_workbook.py"),
    ]
    for k, v in stack:
        info_row(r, k, v); r += 1

    # Apply thin borders for info rows range
    for row_num in range(5, r):
        for col in [2, 3]:
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border()


# ---------------------------------------------------------------------------
# Sheet 2: Master Products
# ---------------------------------------------------------------------------

def build_master_products_sheet(ws, df):
    ws.title = "Master Products"
    ws.sheet_view.showGridLines = False

    col_widths = {
        "A": 12, "B": 14, "C": 38, "D": 22, "E": 12,
        "F": 12, "G": 13, "H": 10, "I": 22, "J": 10, "K": 12, "L": 10,
    }
    num_fmts = {
        "cost_price":    '"$"#,##0.00',
        "selling_price": '"$"#,##0.00',
        "margin_pct":    "0.0",
        "stock_value":   '"$"#,##0.00',
    }

    last_row = df_to_sheet(ws, df, start_row=1, col_widths=col_widths,
                           number_formats=num_fmts)

    # Conditional formatting: Discontinued → light purple highlight on status col
    status_col_idx = list(df.columns).index("status") + 1
    status_col_letter = get_column_letter(status_col_idx)
    range_ref = f"A2:{get_column_letter(len(df.columns))}{last_row + 1}"

    ws.conditional_formatting.add(
        f"{status_col_letter}2:{status_col_letter}{last_row + 1}",
        CellIsRule(operator="equal", formula=['"Discontinued"'],
                   fill=PatternFill(fill_type="solid", fgColor="F3E5F5"),
                   font=Font(color="6C3483", italic=True))
    )

    # Low stock flag → orange highlight
    low_stock_idx = list(df.columns).index("low_stock_flag") + 1
    ls_col = get_column_letter(low_stock_idx)
    ws.conditional_formatting.add(
        f"{ls_col}2:{ls_col}{last_row + 1}",
        CellIsRule(operator="equal", formula=['"YES"'],
                   fill=PatternFill(fill_type="solid", fgColor="FFF3E0"),
                   font=Font(color="E07B39", bold=True))
    )

    # Dropdown validation: status
    status_dv = DataValidation(
        type="list", formula1='"Active,Discontinued,Unknown"',
        allow_blank=True, showDropDown=False
    )
    status_dv.sqref = f"{status_col_letter}2:{status_col_letter}5000"
    ws.add_data_validation(status_dv)

    # Dropdown validation: unit_of_measure
    uom_idx = list(df.columns).index("unit_of_measure") + 1
    uom_col = get_column_letter(uom_idx)
    uom_dv = DataValidation(
        type="list", formula1='"EA,BOX,PACK,ROLL,SET,KIT,L,KG,M,PAIR"',
        allow_blank=True, showDropDown=False
    )
    uom_dv.sqref = f"{uom_col}2:{uom_col}5000"
    ws.add_data_validation(uom_dv)


# ---------------------------------------------------------------------------
# Sheet 3: Vendor Imports
# ---------------------------------------------------------------------------

def build_vendor_imports_sheet(ws, df):
    ws.title = "Vendor Imports"
    ws.sheet_view.showGridLines = False

    col_widths = {
        "A": 12, "B": 12, "C": 38, "D": 22, "E": 12,
        "F": 12, "G": 13, "H": 10, "I": 22, "J": 22, "K": 20, "L": 10,
    }
    num_fmts = {
        "cost_price":    '"$"#,##0.00',
        "selling_price": '"$"#,##0.00',
    }
    df_to_sheet(ws, df, start_row=1, col_widths=col_widths, number_formats=num_fmts)


# ---------------------------------------------------------------------------
# Sheet 4: Duplicate Review
# ---------------------------------------------------------------------------

def build_duplicate_review_sheet(ws, df):
    ws.title = "Duplicate Review"
    ws.sheet_view.showGridLines = False

    if df.empty:
        ws.cell(row=1, column=1, value="No duplicate groups detected.")
        return

    col_widths = {
        "A": 14, "B": 20, "C": 10, "D": 30, "E": 22, "F": 38,
        "G": 22, "H": 12, "I": 13, "J": 10, "K": 14, "L": 28,
    }
    num_fmts = {
        "cost_price":    '"$"#,##0.00',
        "selling_price": '"$"#,##0.00',
    }
    last_row = df_to_sheet(ws, df, start_row=1, col_widths=col_widths,
                           number_formats=num_fmts)

    # Highlight SELECTED rows in light green, DUPLICATE in light orange
    review_col_idx = list(df.columns).index("review_note") + 1
    rev_col = get_column_letter(review_col_idx)

    ws.conditional_formatting.add(
        f"{rev_col}2:{rev_col}{last_row + 1}",
        FormulaRule(
            formula=[f'NOT(ISERROR(SEARCH("SELECTED",{rev_col}2)))'],
            fill=PatternFill(fill_type="solid", fgColor="E8F5E9"),
            font=Font(color="1D8348", bold=True)
        )
    )
    ws.conditional_formatting.add(
        f"{rev_col}2:{rev_col}{last_row + 1}",
        FormulaRule(
            formula=[f'NOT(ISERROR(SEARCH("DUPLICATE",{rev_col}2)))'],
            fill=PatternFill(fill_type="solid", fgColor="FFF3E0"),
            font=Font(color="E07B39")
        )
    )


# ---------------------------------------------------------------------------
# Sheet 5: Data Issues
# ---------------------------------------------------------------------------

def build_data_issues_sheet(ws, df):
    ws.title = "Data Issues"
    ws.sheet_view.showGridLines = False

    if df.empty:
        ws.cell(row=1, column=1, value="No data issues detected.")
        return

    col_widths = {
        "A": 30, "B": 10, "C": 14, "D": 38, "E": 28, "F": 52, "G": 10,
    }
    last_row = df_to_sheet(ws, df, start_row=1, col_widths=col_widths)

    sev_col_idx = list(df.columns).index("severity") + 1
    sev_col = get_column_letter(sev_col_idx)

    ws.conditional_formatting.add(
        f"{sev_col}2:{sev_col}{last_row + 1}",
        CellIsRule(operator="equal", formula=['"HIGH"'],
                   fill=PatternFill(fill_type="solid", fgColor=C_HIGH_BG),
                   font=Font(color=C_HIGH_FG, bold=True))
    )
    ws.conditional_formatting.add(
        f"{sev_col}2:{sev_col}{last_row + 1}",
        CellIsRule(operator="equal", formula=['"MEDIUM"'],
                   fill=PatternFill(fill_type="solid", fgColor=C_MED_BG),
                   font=Font(color=C_MED_FG, bold=True))
    )
    ws.conditional_formatting.add(
        f"{sev_col}2:{sev_col}{last_row + 1}",
        CellIsRule(operator="equal", formula=['"LOW"'],
                   fill=PatternFill(fill_type="solid", fgColor=C_LOW_BG),
                   font=Font(color=C_LOW_FG))
    )

    # Dropdown: resolved (for manual review workflow)
    resolved_col = get_column_letter(len(df.columns) + 1)
    ws.cell(row=1, column=len(df.columns) + 1, value="resolved").font = bold_font(color=C_NAVY)
    ws.cell(row=1, column=len(df.columns) + 1).fill = hex_fill(C_GOLD)
    ws.cell(row=1, column=len(df.columns) + 1).alignment = center_align()
    ws.column_dimensions[resolved_col].width = 12

    resolved_dv = DataValidation(
        type="list", formula1='"Yes,No,Pending"',
        allow_blank=True, showDropDown=False
    )
    resolved_dv.sqref = f"{resolved_col}2:{resolved_col}5000"
    ws.add_data_validation(resolved_dv)


# ---------------------------------------------------------------------------
# Sheet 6: Import Log
# ---------------------------------------------------------------------------

def build_import_log_sheet(ws, df):
    ws.title = "Import Log"
    ws.sheet_view.showGridLines = False

    col_widths = {
        "A": 20, "B": 22, "C": 34, "D": 16, "E": 16, "F": 16, "G": 14, "H": 14, "I": 40,
    }
    df_to_sheet(ws, df, start_row=1, col_widths=col_widths)


# ---------------------------------------------------------------------------
# Sheet 7: Category Summary
# ---------------------------------------------------------------------------

def build_category_summary_sheet(ws, master_df):
    ws.title = "Category Summary"
    ws.sheet_view.showGridLines = False

    # ── By Category ──────────────────────────────────────────────────────────
    cat_grp = master_df.groupby("category", dropna=False).agg(
        total_products=("master_id", "count"),
        active=("status", lambda x: (x == "Active").sum()),
        discontinued=("status", lambda x: (x == "Discontinued").sum()),
        low_stock_count=("low_stock_flag", lambda x: (x == "YES").sum()),
        avg_cost=("cost_price", "mean"),
        avg_selling_price=("selling_price", "mean"),
        total_stock_qty=("stock_qty", "sum"),
        total_stock_value=("stock_value", "sum"),
    ).reset_index()
    cat_grp.columns = [
        "Category", "Total Products", "Active", "Discontinued", "Low Stock",
        "Avg Cost ($)", "Avg Selling Price ($)", "Total Stock Qty", "Total Stock Value ($)"
    ]
    cat_grp = cat_grp.sort_values("Total Products", ascending=False).reset_index(drop=True)

    # ── By Vendor ─────────────────────────────────────────────────────────────
    ven_grp = master_df.groupby("vendor_name", dropna=False).agg(
        total_products=("master_id", "count"),
        active=("status", lambda x: (x == "Active").sum()),
        avg_selling_price=("selling_price", "mean"),
        total_stock_value=("stock_value", "sum"),
    ).reset_index()
    ven_grp.columns = ["Vendor", "Total Products", "Active", "Avg Selling Price ($)", "Total Stock Value ($)"]
    ven_grp = ven_grp.sort_values("Total Products", ascending=False).reset_index(drop=True)

    # ── Status Breakdown ──────────────────────────────────────────────────────
    status_grp = master_df.groupby("status", dropna=False).agg(
        count=("master_id", "count"),
    ).reset_index()
    status_grp.columns = ["Status", "Count"]

    # ── Top 10 by Stock Value ─────────────────────────────────────────────────
    top10 = master_df[master_df["stock_value"].notna()].nlargest(10, "stock_value")[
        ["master_id", "sku", "product_name", "category", "selling_price", "stock_qty", "stock_value"]
    ].copy()
    top10.columns = ["Master ID", "SKU", "Product Name", "Category",
                     "Selling Price ($)", "Stock Qty", "Stock Value ($)"]

    row = 1

    def summary_section(df_sec, title, start_row, hdr_bg=C_BLUE_MID):
        nonlocal row
        # Section header
        ws.row_dimensions[start_row].height = 22
        hdr = ws.cell(row=start_row, column=1, value=title)
        hdr.font      = bold_font(size=12, color=C_WHITE)
        hdr.fill      = hex_fill(hdr_bg)
        hdr.alignment = left_align()
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=len(df_sec.columns))
        end = df_to_sheet(ws, df_sec, start_row=start_row + 1,
                          alt_row_fill=True, freeze_row=False, apply_filter=False)
        return end + 2

    row = summary_section(cat_grp,    "📦  Products by Category",        row, C_NAVY)
    row = summary_section(ven_grp,    "🏭  Products by Vendor",           row, C_BLUE_MID)
    row = summary_section(status_grp, "📊  Status Breakdown",             row, C_TEAL)
    row = summary_section(top10,      "🏆  Top 10 Products by Stock Value", row, C_GOLD)

    # Column widths
    col_w = {"A": 28, "B": 16, "C": 12, "D": 14, "E": 12, "F": 16, "G": 18, "H": 18, "I": 20}
    for col_letter, w in col_w.items():
        ws.column_dimensions[col_letter].width = w

    # Number formats for value columns
    from openpyxl.cell import MergedCell as _MC
    for r in ws.iter_rows():
        for cell in r:
            if isinstance(cell, _MC):
                continue
            if cell.column_letter in ("F", "G", "I"):
                if isinstance(cell.value, float):
                    cell.number_format = '"$"#,##0.00'


# ---------------------------------------------------------------------------
# Sheet 8: Inventory Dashboard
# ---------------------------------------------------------------------------

def build_dashboard_sheet(ws, master_df, dup_df, issues_df):
    ws.title = "Inventory Dashboard"
    ws.sheet_view.showGridLines = False

    # Column widths
    for col_letter in "ABCDEFGHIJKLMN":
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["E"].width = 2
    ws.column_dimensions["I"].width = 2

    # ── KPI calculations ──────────────────────────────────────────────────────
    total_products  = len(master_df)
    active_ct       = int((master_df["status"] == "Active").sum())
    disc_ct         = int((master_df["status"] == "Discontinued").sum())
    dup_groups      = master_df["sku"].nunique() if "sku" in master_df.columns else 0
    # count SKUs that appear in dup_df
    dup_group_ct    = int(dup_df["sku"].nunique()) if not dup_df.empty else 0
    total_issues    = len(issues_df)
    high_issues     = int((issues_df["severity"] == "HIGH").sum()) if not issues_df.empty else 0
    low_stock       = int((master_df["low_stock_flag"] == "YES").sum())
    inv_value       = master_df["stock_value"].sum()
    inv_value_str   = f"${inv_value:,.0f}" if not pd.isna(inv_value) else "N/A"

    # ── Title Banner ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 46
    ws.row_dimensions[3].height = 24

    title = ws.cell(row=2, column=2, value="NEXA PARTS SUPPLY")
    title.font      = Font(name="Calibri", bold=True, size=26, color=C_GOLD)
    title.fill      = hex_fill(C_NAVY)
    title.alignment = left_align()

    subtitle = ws.cell(row=2, column=6, value="Inventory Dashboard")
    subtitle.font      = Font(name="Calibri", size=14, color=C_LIGHT_GRAY)
    subtitle.fill      = hex_fill(C_NAVY)
    subtitle.alignment = left_align()

    gen_cell = ws.cell(row=3, column=2,
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}   |   Portfolio Demo — Synthetic Data Only")
    gen_cell.font      = Font(name="Calibri", size=9, color=C_DARK_GRAY, italic=True)
    gen_cell.alignment = left_align()

    # Extend navy fill across banner columns
    for c in range(2, 14):
        ws.cell(row=2, column=c).fill = hex_fill(C_NAVY)

    # ── KPI Card helper ───────────────────────────────────────────────────────
    def kpi_card(start_row, start_col, label, value, icon="", bg=C_BLUE_MID, value_color=C_WHITE):
        # Card background: 3 rows × 3 cols
        card_rows = [start_row, start_row + 1, start_row + 2]
        card_cols = [start_col, start_col + 1, start_col + 2]

        for r in card_rows:
            ws.row_dimensions[r].height = 20
            for c in card_cols:
                cell = ws.cell(row=r, column=c)
                cell.fill = hex_fill(bg)

        # Icon + label
        label_cell = ws.cell(row=start_row, column=start_col, value=f"{icon}  {label}")
        label_cell.font      = Font(name="Calibri", size=9, color="CCDDEE", bold=False)
        label_cell.fill      = hex_fill(bg)
        label_cell.alignment = left_align()

        # Value (big, centred across 3 cols)
        val_cell = ws.cell(row=start_row + 1, column=start_col, value=value)
        val_cell.font      = Font(name="Calibri", bold=True, size=22, color=value_color)
        val_cell.fill      = hex_fill(bg)
        val_cell.alignment = center_align()
        ws.merge_cells(start_row=start_row + 1, start_column=start_col,
                       end_row=start_row + 1, end_column=start_col + 2)

        # Bottom accent line
        bottom_fill = PatternFill(fill_type="solid", fgColor=C_GOLD)
        for c in card_cols:
            ws.cell(row=start_row + 2, column=c).fill = bottom_fill
            ws.row_dimensions[start_row + 2].height = 4

    # ── KPI Row 1 ─────────────────────────────────────────────────────────────
    ws.row_dimensions[5].height = 8   # spacer
    kpi_card(6, 2,  "Total Products",    total_products,    "📦", C_NAVY)
    kpi_card(6, 6,  "Active Products",   active_ct,         "✅", C_TEAL)
    kpi_card(6, 10, "Discontinued",      disc_ct,           "🚫", "4A4A6A")

    # ── KPI Row 2 ─────────────────────────────────────────────────────────────
    ws.row_dimensions[10].height = 8
    kpi_card(11, 2,  "Duplicate Groups",  dup_group_ct,       "🔁", C_BLUE_MID)
    kpi_card(11, 6,  "Total Issues",      total_issues,       "⚠", C_ORANGE)
    kpi_card(11, 10, "High Severity",     high_issues,        "🔴", C_RED)

    # ── KPI Row 3 ─────────────────────────────────────────────────────────────
    ws.row_dimensions[15].height = 8
    kpi_card(16, 2,  "Low-Stock Items",   low_stock,          "📉", "D4860E")
    kpi_card(16, 6,  "Est. Inv. Value",   inv_value_str,      "💰", C_GREEN, C_WHITE)

    # Spacer
    ws.row_dimensions[20].height = 16

    # ── Summary Tables ────────────────────────────────────────────────────────
    # Category summary mini-table
    cat_summary = master_df.groupby("category").agg(
        products=("master_id", "count"),
        active=("status", lambda x: (x == "Active").sum()),
        stock_value=("stock_value", "sum"),
    ).reset_index().sort_values("products", ascending=False).head(12)
    cat_summary.columns = ["Category", "Products", "Active", "Stock Value ($)"]

    row_start = 21
    tbl_title = ws.cell(row=row_start, column=2, value="📦  Products by Category")
    tbl_title.font      = bold_font(size=11, color=C_WHITE)
    tbl_title.fill      = hex_fill(C_NAVY)
    tbl_title.alignment = left_align()
    ws.merge_cells(start_row=row_start, start_column=2, end_row=row_start, end_column=5)

    end_cat = df_to_sheet(ws, cat_summary, start_row=row_start + 1,
                          alt_row_fill=True, freeze_row=False, apply_filter=False)
    for r in ws.iter_rows(min_row=row_start + 2, max_row=end_cat, min_col=5, max_col=5):
        for cell in r:
            if isinstance(cell.value, float):
                cell.number_format = '"$"#,##0.00'

    # Vendor summary mini-table
    ven_summary = master_df.groupby("vendor_name").agg(
        products=("master_id", "count"),
        active=("status", lambda x: (x == "Active").sum()),
        stock_value=("stock_value", "sum"),
    ).reset_index().sort_values("products", ascending=False)
    ven_summary.columns = ["Vendor", "Products", "Active", "Stock Value ($)"]

    ven_row_start = 21
    tbl_ven_title = ws.cell(row=ven_row_start, column=8, value="🏭  Products by Vendor")
    tbl_ven_title.font      = bold_font(size=11, color=C_WHITE)
    tbl_ven_title.fill      = hex_fill(C_BLUE_MID)
    tbl_ven_title.alignment = left_align()
    ws.merge_cells(start_row=ven_row_start, start_column=8, end_row=ven_row_start, end_column=11)

    end_ven = df_to_sheet(ws, ven_summary, start_row=ven_row_start + 1,
                          alt_row_fill=True, freeze_row=False, apply_filter=False,
                          col_widths=None)
    # Manually position vendor table columns (it writes from col 1 by default with df_to_sheet)
    # We need to re-write it starting at col 8
    # Clear previous write and re-do with offset
    # Better approach: write vendor table manually at offset
    start_c = 8
    headers_v = list(ven_summary.columns)
    apply_header_row(ws, ven_row_start + 1, headers_v, col_start=start_c, bg=C_BLUE_MID)
    for r_idx, row_data in enumerate(ven_summary.itertuples(index=False), start=ven_row_start + 2):
        is_alt = (r_idx - ven_row_start) % 2 == 0
        rfill = hex_fill(C_ROW_ALT) if is_alt else hex_fill(C_ROW_NORM)
        for c_idx, val in enumerate(row_data, start=start_c):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font      = normal_font()
            cell.fill      = rfill
            cell.alignment = left_align()
            cell.border    = thin_border()
            if c_idx == start_c + 3 and isinstance(val, float):
                cell.number_format = '"$"#,##0.00'

    # Issue severity summary
    if not issues_df.empty:
        sev_summary = issues_df.groupby("severity").size().reset_index(name="Count")
        sev_row = end_cat + 2

        tbl_sev_title = ws.cell(row=sev_row, column=2, value="⚠  Issue Severity Breakdown")
        tbl_sev_title.font      = bold_font(size=11, color=C_WHITE)
        tbl_sev_title.fill      = hex_fill(C_ORANGE)
        tbl_sev_title.alignment = left_align()
        ws.merge_cells(start_row=sev_row, start_column=2, end_row=sev_row, end_column=5)

        apply_header_row(ws, sev_row + 1, ["Severity", "Count"], col_start=2, bg=C_ORANGE)
        for r_idx, sev_row_data in enumerate(sev_summary.itertuples(index=False), start=sev_row + 2):
            sev_val = sev_row_data[0]
            cnt_val = sev_row_data[1]
            bg_map  = {"HIGH": C_HIGH_BG, "MEDIUM": C_MED_BG, "LOW": C_LOW_BG}
            fg_map  = {"HIGH": C_HIGH_FG, "MEDIUM": C_MED_FG, "LOW": C_LOW_FG}
            rfill   = hex_fill(bg_map.get(sev_val, C_WHITE))
            sev_cell = ws.cell(row=r_idx, column=2, value=sev_val)
            sev_cell.font  = Font(name="Calibri", bold=True, color=fg_map.get(sev_val, C_BLACK))
            sev_cell.fill  = rfill
            sev_cell.border = thin_border()
            cnt_cell = ws.cell(row=r_idx, column=3, value=cnt_val)
            cnt_cell.font  = normal_font()
            cnt_cell.fill  = rfill
            cnt_cell.border = thin_border()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 62)
    print("  Nexa Parts Supply - Workbook Builder")
    print("  Portfolio Demo | Synthetic Data Only")
    print("=" * 62)

    # Load processed data
    def load(path, label):
        if not path.exists():
            print(f"  [ERR] {path.name} not found - run clean_and_merge.py first")
            import sys; sys.exit(1)
        df = pd.read_csv(path)
        print(f"  [OK] {path.name:<42} {len(df):>4} rows  -> {label}")
        return df

    print("\n  Loading processed files...")
    master_df   = load(PROC_DIR / "master_products.csv",  "Master Products")
    imports_df  = load(PROC_DIR / "vendor_imports.csv",   "Vendor Imports")
    dup_df      = load(RPT_DIR  / "duplicate_review.csv", "Duplicate Review")
    issues_df   = load(RPT_DIR  / "data_issues.csv",      "Data Issues")
    log_df      = load(RPT_DIR  / "import_log.csv",       "Import Log")

    print("\n  Building workbook sheets...")
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    sheet_builders = [
        ("README",              lambda ws: build_readme_sheet(ws)),
        ("Master Products",     lambda ws: build_master_products_sheet(ws, master_df)),
        ("Vendor Imports",      lambda ws: build_vendor_imports_sheet(ws, imports_df)),
        ("Duplicate Review",    lambda ws: build_duplicate_review_sheet(ws, dup_df)),
        ("Data Issues",         lambda ws: build_data_issues_sheet(ws, issues_df)),
        ("Import Log",          lambda ws: build_import_log_sheet(ws, log_df)),
        ("Category Summary",    lambda ws: build_category_summary_sheet(ws, master_df)),
        ("Inventory Dashboard", lambda ws: build_dashboard_sheet(ws, master_df, dup_df, issues_df)),
    ]

    for sheet_name, builder in sheet_builders:
        ws = wb.create_sheet(title=sheet_name)
        builder(ws)
        print(f"  [OK] Sheet: {sheet_name}")

    wb.save(OUT_FILE)

    print(f"\n  " + "-" * 53)
    print(f"  Workbook saved: {OUT_FILE}")
    print(f"  Sheets        : {len(wb.sheetnames)}")
    print("  " + "-" * 53)
    print("=" * 62)


if __name__ == "__main__":
    main()
